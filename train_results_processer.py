""" This script  process results from YOLO training to overview table. 
"""

import subprocess
import os
import sys
sys.path.append(r"C:\\Users\\ingrj\\AppData\\Roaming\\Python\\Python312\\site-packages")
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


SOURCE_PATH = "runs/final_m_default"
CREATE_CHART = True
PARAMETER_NAME = "dataset"
X_AXIS_NAME = "partial datasets"
TITLE_PARAM_NAME = "parciálních datasetech"
X_AXIS_SCALE = '' # default = linear

def load_training_results(train_folder : str) -> dict:
    """ Load training results folder and return dictionary with mAP50
        and mAP50:95 for each epoch. """
    train_path =  os.path.join(SOURCE_PATH, train_folder, "results.csv")
    df = pd.read_csv(train_path)
    results = {
        "map50" : df["metrics/mAP50(B)"],
        "map50_95" : df["metrics/mAP50-95(B)"],    
    }
    return results

def create_excel(all_trainings : dict) -> None:
    """ Create excel table with comparsion of all trainings in source folder. """
    max_epochs = max(len(one_train["map50"]) for one_train in all_trainings.values())
    df_dict = {'epocha' : [""] + list(range(1, max_epochs + 1)) + ["-"] + ["MAX"]}

    count = 0
    for name, data in all_trainings.items():
        column_letter_even = get_column_letter(count + 2)
        column_letter_odd = get_column_letter(count + 3)
        map50 = (["mAP50"]
                 + list(data["map50"])
                 + [None] * (max_epochs - len(data["map50"]))
                 + [""]
                 + [f"=MAX({column_letter_even}2:{column_letter_even}{max_epochs+1})"])
        map50_95 = (["mAP50:95"]
                    + list(data["map50_95"])
                    + [None] * (max_epochs - len(data["map50_95"]))
                    + [""]
                    + [f"=MAX({column_letter_odd}2:{column_letter_odd}{max_epochs+1})"])
        df_dict[name] = map50
        df_dict[f"{name}_map50_95"] = map50_95
        count += 2

    df = pd.DataFrame(df_dict)

    writer = pd.ExcelWriter(os.path.join(SOURCE_PATH, "overview.xlsx"), engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="Overview")

    wb = writer.book
    ws = wb["Overview"]

    for col_idx in range(2, len(all_trainings) * 2 + 2, 2):
        col_letter1 = get_column_letter(col_idx)
        col_letter2 = get_column_letter(col_idx + 1)
        ws.merge_cells(f"{col_letter1}1:{col_letter2}1")
        ws[f"{col_letter1}1"].alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    writer._save()

# load results
all_train_res = {}
result_folders = list(train for train in os.listdir(SOURCE_PATH)
                       if os.path.isdir(os.path.join(SOURCE_PATH, train)))
for train_folder_name in result_folders:
    train_results = load_training_results(train_folder_name)
    all_train_res[train_folder_name] = train_results

# crate overview excel table
create_excel(all_train_res)
print("Overview file was created successfully!")

if CREATE_CHART:
    subprocess.run(["python", "bar_plot_module.py", SOURCE_PATH, 
                    PARAMETER_NAME, X_AXIS_NAME, X_AXIS_SCALE, TITLE_PARAM_NAME])